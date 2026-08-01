import os
import shutil
import uuid
import secrets
from datetime import datetime
from typing import List, Optional
import pandas as pd
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

import config, database, models, email_service
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Initialize database on startup
database.init_db()

app = FastAPI(
    title="DPGU STR Induction Management API",
    description="Backend API for student registration, QR code generation, check-in, and stats reporting.",
    version="1.0.0"
)

# CORS Setup
app.add_middleware(
    CORSMiddleware,
    allow_origins=config.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create static folder and subfolders if not present, and mount them
os.makedirs("static", exist_ok=True)
os.makedirs(config.QRCODES_DIR, exist_ok=True)
os.makedirs(config.MOCK_EMAILS_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def read_root():
    return {
        "app": "DPGU STR Induction Management System",
        "status": "Online",
        "timestamp": datetime.now().isoformat()
    }

@app.post("/api/students/upload", response_model=models.StatsResponse)
def upload_students(
    file: UploadFile = File(...),
    dept: Optional[str] = Query(None)
):
    """Upload Excel/CSV file of students and save them to the database."""
    # Verify file extension
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".csv", ".xlsx", ".xls"]:
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload CSV or Excel.")

    # Save uploaded file temporarily
    temp_path = os.path.join(config.UPLOAD_DIR, f"{uuid.uuid4()}{ext}")
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Read using Pandas
        if ext == ".csv":
            df = pd.read_csv(temp_path)
        else:
            df = pd.read_excel(temp_path)
            
        # Clean columns (lowercase, strip whitespace)
        df.columns = [c.lower().strip() for c in df.columns]
        
        # Map columns flexibly
        name_col = None
        email_col = None
        dept_col = None
        
        for col in df.columns:
            if "name" in col:
                name_col = col
            elif "email" in col or "mail" in col:
                email_col = col
            elif "dept" in col or "department" in col or "stream" in col:
                dept_col = col
                
        if not name_col or not email_col:
            raise HTTPException(
                status_code=400, 
                detail="CSV/Excel must contain at least 'Name' and 'Email' columns."
            )
            
        # Insert records in SQLite
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Fetch current maximum id to keep tokens sequential
        res = cursor.execute("SELECT COUNT(*) FROM students").fetchone()
        count = res[0] if res else 0
        
        inserted_count = 0
        for index, row in df.iterrows():
            name = str(row[name_col]).strip()
            email = str(row[email_col]).strip()
            department = str(row[dept_col]).strip() if dept_col and pd.notna(row[dept_col]) else (dept or "General")
            
            # Simple skip for empty rows
            if not name or not email or name.lower() == "nan" or email.lower() == "nan":
                continue
                
            # Check if student email already exists to avoid duplicates
            existing = cursor.execute("SELECT id FROM students WHERE email = ?", (email,)).fetchone()
            if existing:
                continue
                
            # Generate unique token e.g., DPGU-IND-1001-C7F3
            token_num = 1001 + count + inserted_count
            token_suffix = secrets.token_hex(2).upper()
            student_id = f"DPGU-IND-{token_num}-{token_suffix}"
            
            cursor.execute(
                """
                INSERT INTO students (student_id, name, email, department)
                VALUES (?, ?, ?, ?)
                """,
                (student_id, name, email, department)
            )
            inserted_count += 1
            
        conn.commit()
        conn.close()
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
    finally:
        # Cleanup temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
    return get_stats(dept)

@app.get("/api/students", response_model=List[models.StudentResponse])
def get_students(
    search: Optional[str] = Query(None, description="Search by name, email, or department"),
    checked_in: Optional[bool] = Query(None, description="Filter by check-in status"),
    email_sent: Optional[bool] = Query(None, description="Filter by email sent status"),
    dept: Optional[str] = Query(None)
):
    """Retrieve all students with optional search filters."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    
    query = "SELECT * FROM students WHERE 1=1"
    params = []
    
    if search:
        query += " AND (name LIKE ? OR email LIKE ? OR department LIKE ?)"
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])
        
    if checked_in is not None:
        query += " AND checked_in = ?"
        params.append(True if checked_in else False)
        
    if email_sent is not None:
        query += " AND email_sent = ?"
        params.append(True if email_sent else False)
        
    if dept:
        query += " AND LOWER(department) = LOWER(?)"
        params.append(dept)
        
    query += " ORDER BY id DESC"
    
    rows = cursor.execute(query, params).fetchall()
    conn.close()
    
    students = []
    for r in rows:
        students.append(
            models.StudentResponse(
                id=r["id"],
                student_id=r["student_id"],
                name=r["name"],
                email=r["email"],
                department=r["department"],
                checked_in=bool(r["checked_in"]),
                check_in_time=r["check_in_time"],
                email_sent=bool(r["email_sent"]),
                email_sent_time=r["email_sent_time"]
            )
        )
    return students

@app.post("/api/students/send-emails")
async def send_student_emails(
    background_tasks: BackgroundTasks,
    dept: Optional[str] = Query(None)
):
    """Trigger email generation and delivery in the background for students who haven't received them yet."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    # Fetch all students who haven't received their email yet
    if dept:
        rows = cursor.execute("SELECT * FROM students WHERE email_sent = FALSE AND LOWER(department) = LOWER(?)", (dept,)).fetchall()
    else:
        rows = cursor.execute("SELECT * FROM students WHERE email_sent = FALSE").fetchall()
    conn.close()
    
    if not rows:
        return {"message": "All students have already received their emails. No new emails to send."}
        
    # Load the persisted SMTP / Mock settings from SQLite/Postgres database
    active_settings = database.get_settings(dept)
    
    def process_email_queue(students_list, config_override):
        db_conn = database.get_db_connection()
        db_cursor = db_conn.cursor()
        for s in students_list:
            success = email_service.send_email(
                student_name=s["name"],
                student_email=s["email"],
                student_id=s["student_id"],
                student_department=s.get("department", "General"),
                smtp_settings=config_override
            )
            if success:
                db_cursor.execute(
                    """
                    UPDATE students
                    SET email_sent = TRUE, email_sent_time = ?
                    WHERE id = ?
                    """,
                    (datetime.now().isoformat(), s["id"])
                )
                db_conn.commit()
        db_conn.close()
        
    background_tasks.add_task(process_email_queue, [dict(r) for r in rows], active_settings)
    
    return {"message": f"Queued {len(rows)} emails to send in the background."}

@app.get("/api/settings", response_model=models.EmailConfigRequest)
def get_settings(dept: Optional[str] = Query(None)):
    """Retrieve active email configuration settings from database."""
    settings = database.get_settings(dept)
    if not settings:
        raise HTTPException(status_code=404, detail="Settings not found.")
    return models.EmailConfigRequest(
        mock_email=bool(settings["mock_email"]),
        smtp_host=settings["smtp_host"] or "",
        smtp_port=settings["smtp_port"] or 587,
        smtp_user=settings["smtp_user"] or "",
        smtp_password=settings["smtp_password"] or "",
        smtp_from=settings["smtp_from"] or "",
        smtp_from_name=settings["smtp_from_name"] or "",
        email_subject=settings.get("email_subject") or "Your DPGU STR Induction Pass & Invitation",
        email_body=settings.get("email_body") or "",
        event_date=settings.get("event_date") or "August 5, 2026",
        event_time=settings.get("event_time") or "09:00 AM",
        event_venue=settings.get("event_venue") or "Main Auditorium"
    )

@app.post("/api/settings")
def save_settings(settings: models.EmailConfigRequest, dept: Optional[str] = Query(None)):
    """Save email configuration settings to database."""
    database.update_settings(settings.model_dump(), dept)
    return {"message": "Settings saved successfully."}

@app.post("/api/students/checkin", response_model=models.CheckInResponse)
def checkin_student(request: models.StudentCheckIn, dept: Optional[str] = Query(None)):
    """Mark a student as checked in after scanning their QR code, enforcing desk-specific validation."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    
    student = cursor.execute("SELECT * FROM students WHERE student_id = ? FOR UPDATE", (request.student_id,)).fetchone()
    
    if not student:
        conn.close()
        return models.CheckInResponse(
            success=False,
            message="Invalid QR Code. Student not found in database."
        )
        
    student_data = models.StudentResponse(
        id=student["id"],
        student_id=student["student_id"],
        name=student["name"],
        email=student["email"],
        department=student["department"],
        checked_in=bool(student["checked_in"]),
        check_in_time=student["check_in_time"],
        email_sent=bool(student["email_sent"]),
        email_sent_time=student["email_sent_time"]
    )
    
    if dept and student["department"] and student["department"].strip().lower() != dept.strip().lower():
        conn.close()
        return models.CheckInResponse(
            success=False,
            message=f"Student belongs to {student['department']} department. Please refer them to the {student['department']} desk.",
            student=student_data
        )
        
    if student["checked_in"]:
        conn.close()
        return models.CheckInResponse(
            success=False,
            message=f"{student['name']} is already checked in at {student['check_in_time']}.",
            student=student_data
        )
        
    check_in_time = datetime.now().isoformat()
    cursor.execute(
        """
        UPDATE students
        SET checked_in = TRUE, check_in_time = ?
        WHERE student_id = ?
        """,
        (check_in_time, request.student_id)
    )
    conn.commit()
    
    updated_student = cursor.execute("SELECT * FROM students WHERE student_id = ?", (request.student_id,)).fetchone()
    conn.close()
    
    student_data = models.StudentResponse(
        id=updated_student["id"],
        student_id=updated_student["student_id"],
        name=updated_student["name"],
        email=updated_student["email"],
        department=updated_student["department"],
        checked_in=bool(updated_student["checked_in"]),
        check_in_time=updated_student["check_in_time"],
        email_sent=bool(updated_student["email_sent"]),
        email_sent_time=updated_student["email_sent_time"]
    )
    
    return models.CheckInResponse(
        success=True,
        message=f"Welcome {updated_student['name']}! Attendance marked successfully.",
        student=student_data
    )

@app.post("/api/students/checkout", response_model=models.CheckInResponse)
def checkout_student(request: models.StudentCheckIn, dept: Optional[str] = Query(None)):
    """Revert check-in status (mark as checked out/pending) for manual admin corrections."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    
    student = cursor.execute("SELECT * FROM students WHERE student_id = ? FOR UPDATE", (request.student_id,)).fetchone()
    
    if not student:
        conn.close()
        return models.CheckInResponse(
            success=False,
            message="Student not found."
        )
        
    student_data = models.StudentResponse(
        id=student["id"],
        student_id=student["student_id"],
        name=student["name"],
        email=student["email"],
        department=student["department"],
        checked_in=bool(student["checked_in"]),
        check_in_time=student["check_in_time"],
        email_sent=bool(student["email_sent"]),
        email_sent_time=student["email_sent_time"]
    )
    
    if dept and student["department"] and student["department"].strip().lower() != dept.strip().lower():
        conn.close()
        return models.CheckInResponse(
            success=False,
            message=f"Student belongs to {student['department']} department. Cannot revert check-in from {dept} desk.",
            student=student_data
        )
        
    cursor.execute(
        """
        UPDATE students
        SET checked_in = FALSE, check_in_time = NULL
        WHERE student_id = ?
        """,
        (request.student_id,)
    )
    conn.commit()
    
    updated_student = cursor.execute("SELECT * FROM students WHERE student_id = ?", (request.student_id,)).fetchone()
    conn.close()
    
    student_data = models.StudentResponse(
        id=updated_student["id"],
        student_id=updated_student["student_id"],
        name=updated_student["name"],
        email=updated_student["email"],
        department=updated_student["department"],
        checked_in=bool(updated_student["checked_in"]),
        check_in_time=updated_student["check_in_time"],
        email_sent=bool(updated_student["email_sent"]),
        email_sent_time=updated_student["email_sent_time"]
    )
    
    return models.CheckInResponse(
        success=True,
        message=f"Checked out {updated_student['name']} successfully.",
        student=student_data
    )

@app.get("/api/stats", response_model=models.StatsResponse)
def get_stats_endpoint(dept: Optional[str] = Query(None)):
    """Fetch current system statistics."""
    return get_stats(dept)

def get_stats(dept: Optional[str] = None) -> models.StatsResponse:
    conn = database.get_db_connection()
    cursor = conn.cursor()
    
    if dept:
        total = cursor.execute("SELECT COUNT(*) FROM students WHERE LOWER(department) = LOWER(?)", (dept,)).fetchone()[0]
        checked_in = cursor.execute("SELECT COUNT(*) FROM students WHERE checked_in = TRUE AND LOWER(department) = LOWER(?)", (dept,)).fetchone()[0]
        emails_sent = cursor.execute("SELECT COUNT(*) FROM students WHERE email_sent = TRUE AND LOWER(department) = LOWER(?)", (dept,)).fetchone()[0]
    else:
        total = cursor.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        checked_in = cursor.execute("SELECT COUNT(*) FROM students WHERE checked_in = TRUE").fetchone()[0]
        emails_sent = cursor.execute("SELECT COUNT(*) FROM students WHERE email_sent = TRUE").fetchone()[0]
    
    conn.close()
    
    rate = 0.0
    if total > 0:
        rate = round((checked_in / total) * 100, 2)
        
    return models.StatsResponse(
        total_students=total,
        checked_in=checked_in,
        emails_sent=emails_sent,
        check_in_rate=rate
    )

@app.get("/api/students/export")
def export_students(dept: Optional[str] = Query(None)):
    """Export attendance list as a styled and protected Excel (.xlsx) file."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    if dept:
        rows = cursor.execute("SELECT student_id, name, email, department, checked_in, check_in_time, email_sent, email_sent_time FROM students WHERE LOWER(department) = LOWER(?)", (dept,)).fetchall()
    else:
        rows = cursor.execute("SELECT student_id, name, email, department, checked_in, check_in_time, email_sent, email_sent_time FROM students").fetchall()
    conn.close()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Induction Attendance 2026"
    
    # Style definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # DPGU Indigo
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    checked_in_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light green
    pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light orange/yellow
    
    body_font = Font(name=font_family, size=10, bold=False, color="000000")
    bold_body_font = Font(name=font_family, size=10, bold=True, color="000000")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Write Headers
    headers = [
        "Student ID", "Full Name", "Email Address", "Department", 
        "Check-In Status", "Check-In Time", "Invitation Status", "Sent Time"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Write Rows
    for row_idx, r in enumerate(rows, 2):
        # Data values
        student_id = r["student_id"]
        name = r["name"]
        email = r["email"]
        dept_name = r["department"] or "General"
        
        checked_in = "Checked In" if bool(r["checked_in"]) else "Pending"
        check_in_time = r["check_in_time"] or "-"
        
        email_sent = "Emailed" if bool(r["email_sent"]) else "Pending"
        email_sent_time = r["email_sent_time"] or "-"
        
        row_data = [
            student_id, name, email, dept_name, 
            checked_in, check_in_time, email_sent, email_sent_time
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = left_align if col_idx in [2, 3, 4] else center_align
            
            # Highlight token
            if col_idx == 1:
                cell.font = bold_body_font
                
            # Formatting for Status column
            if col_idx == 5:
                if val == "Checked In":
                    cell.fill = checked_in_fill
                    cell.font = Font(name=font_family, size=10, bold=True, color="047857")
                else:
                    cell.fill = pending_fill
                    cell.font = Font(name=font_family, size=10, color="B45309")
            
            # Formatting for Email column
            if col_idx == 7:
                if val == "Emailed":
                    cell.font = Font(name=font_family, size=10, bold=True, color="4F46E5")
                else:
                    cell.font = Font(name=font_family, size=10, color="6B7280")
                    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Set Row Heights
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 22
        
    # Enable sheet protection
    ws.protection.sheet = True
    ws.protection.password = "DPGU-STR-2026"
    ws.protection.enable()
    
    # Save file
    export_path = os.path.join(config.UPLOAD_DIR, "induction_attendance_report.xlsx")
    wb.save(export_path)
    
    return FileResponse(
        export_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"DPGU_Induction_Attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

@app.post("/api/students/reset")
def reset_system(dept: Optional[str] = Query(None)):
    """Clear all records and delete generated static files, optionally filtered by department."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    ph = "%s" if config.IS_POSTGRES else "?"
    
    if dept:
        # Fetch all students in this department to get their student_ids and emails
        students = cursor.execute(f"SELECT student_id, email FROM students WHERE LOWER(department) = LOWER({ph})", (dept,)).fetchall()
        for s in students:
            # Delete QR codes
            qr_file = os.path.join(config.QRCODES_DIR, f"{s['student_id']}.png")
            if os.path.exists(qr_file):
                try:
                    os.unlink(qr_file)
                except Exception:
                    pass
            # Delete mock email previews
            preview_file = os.path.join(config.MOCK_EMAILS_DIR, f"{s['email']}_{s['student_id']}.png")
            if os.path.exists(preview_file):
                try:
                    os.unlink(preview_file)
                except Exception:
                    pass
            # Check for failed email previews
            failed_preview_file = os.path.join(config.MOCK_EMAILS_DIR, f"FAILED_SMTP_{s['email']}_{s['student_id']}.png")
            if os.path.exists(failed_preview_file):
                try:
                    os.unlink(failed_preview_file)
                except Exception:
                    pass
                    
        # Delete from students table
        cursor.execute(f"DELETE FROM students WHERE LOWER(department) = LOWER({ph})", (dept,))
        # Delete department-specific settings row
        cursor.execute(f"DELETE FROM settings WHERE LOWER(department) = LOWER({ph})", (dept,))
        conn.commit()
        conn.close()
        return {"message": f"Wiped all records and passes for department '{dept}'."}
    else:
        conn.close()
        database.clear_db()
        
        # Clean folders
        for folder in [config.QRCODES_DIR, config.MOCK_EMAILS_DIR]:
            if os.path.exists(folder):
                for filename in os.listdir(folder):
                    file_path = os.path.join(folder, filename)
                    try:
                        if os.path.isfile(file_path) or os.path.islink(file_path):
                            os.unlink(file_path)
                    except Exception as e:
                        print(f"Failed to delete {file_path}. Reason: {e}")
                        
        return {"message": "Database wiped and all generated files deleted successfully."}

@app.get("/api/preview-emails")
def get_mock_emails(dept: Optional[str] = Query(None)):
    """Return list of generated PNG passes for preview, optionally filtered by department."""
    emails = []
    
    dept_student_ids = set()
    if dept:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        rows = cursor.execute("SELECT student_id FROM students WHERE LOWER(department) = LOWER(?)", (dept,)).fetchall()
        conn.close()
        dept_student_ids = {r["student_id"] for r in rows}
        
    if os.path.exists(config.MOCK_EMAILS_DIR):
        for f in os.listdir(config.MOCK_EMAILS_DIR):
            if f.endswith(".png"):
                name_without_ext = f[:-4]  # Strip ".png"
                if name_without_ext.startswith("FAILED_SMTP_"):
                    content = name_without_ext[12:]
                    email_part, id_part = content.rsplit("_", 1)
                elif name_without_ext.startswith("FAILED_"):
                    content = name_without_ext[7:]
                    email_part, id_part = content.rsplit("_", 1)
                else:
                    email_part, id_part = name_without_ext.rsplit("_", 1)
                
                if dept and id_part not in dept_student_ids:
                    continue
                    
                emails.append({
                    "filename": f,
                    "email": email_part,
                    "student_id": id_part,
                    "file_url": f"/static/sent_emails/{f}"
                })
    return emails

@app.post("/api/preview-emails/clear")
def clear_mock_emails(dept: Optional[str] = Query(None)):
    """Clear generated PNG email passes, optionally filtered by department."""
    if not os.path.exists(config.MOCK_EMAILS_DIR):
        return {"message": "Previews folder empty."}
        
    dept_student_ids = set()
    if dept:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        rows = cursor.execute("SELECT student_id FROM students WHERE LOWER(department) = LOWER(?)", (dept,)).fetchall()
        conn.close()
        dept_student_ids = {r["student_id"] for r in rows}
        
    for filename in os.listdir(config.MOCK_EMAILS_DIR):
        if filename.endswith(".png"):
            if dept:
                name_without_ext = filename[:-4]
                try:
                    if "_" in name_without_ext:
                        id_part = name_without_ext.split("_")[-1]
                        if id_part not in dept_student_ids:
                            continue
                except Exception:
                    continue
                    
            file_path = os.path.join(config.MOCK_EMAILS_DIR, filename)
            try:
                os.unlink(file_path)
            except Exception:
                pass
                
    return {"message": "Mock email previews cleared successfully."}

@app.post("/api/settings/test-smtp")
def test_smtp_connection(settings: models.EmailConfigRequest):
    """Verify SMTP configuration by sending a test email."""
    if settings.mock_email:
        return {"success": True, "message": "Dry-run/Mock Mode is enabled. No SMTP request made."}
        
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "SMTP Test Connection - DPGU STR Induction"
        msg["From"] = f"{settings.smtp_from_name} <{settings.smtp_from}>"
        msg["To"] = settings.smtp_user  # Send test email to the user account itself
        
        body = "<h3>SMTP Test Successful!</h3><p>Your DPGU STR Induction email setup is fully operational.</p>"
        msg.attach(MIMEText(body, "html"))
        
        server = smtplib.SMTP(settings.smtp_host, settings.smtp_port)
        server.starttls()
        server.login(settings.smtp_user, settings.smtp_password)
        server.sendmail(settings.smtp_from, settings.smtp_user, msg.as_string())
        server.quit()
        return {"success": True, "message": f"Test email sent successfully to {settings.smtp_user}."}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.get("/api/auth/profiles", response_model=List[models.UserResponse])
def get_auth_profiles():
    """Retrieve all user profiles for login selection."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    rows = cursor.execute("SELECT id, username, name, role, department, is_authorized FROM users ORDER BY role ASC, name ASC").fetchall()
    conn.close()
    return [dict(row) for row in rows]

@app.post("/api/auth/login", response_model=models.UserResponse)
def login(payload: models.UserLoginRequest):
    """Authenticate a user profile using a 4-digit PIN."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    row = cursor.execute("SELECT * FROM users WHERE username = ?", (payload.username,)).fetchone()
    conn.close()
    
    if not row:
        raise HTTPException(status_code=401, detail="Invalid username or PIN.")
        
    user = dict(row)
    if user["pin"] != payload.pin:
        raise HTTPException(status_code=401, detail="Incorrect PIN.")
        
    if not user["is_authorized"]:
        raise HTTPException(status_code=403, detail="Your account access has been suspended or unauthorized by the administrator.")
        
    return user

@app.get("/api/users", response_model=List[models.UserResponse])
def get_users(dept: Optional[str] = Query(None)):
    """List all pre-configured employee users."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    if dept:
        rows = cursor.execute("SELECT * FROM users WHERE role = 'employee' AND LOWER(department) = LOWER(?)", (dept,)).fetchall()
    else:
        rows = cursor.execute("SELECT * FROM users WHERE role = 'employee'").fetchall()
    conn.close()
    return [dict(row) for row in rows]

@app.post("/api/users/{username}/toggle-auth")
def toggle_user_auth(username: str):
    """Grant or revoke access authorization for a user."""
    conn = database.get_db_connection()
    cursor = conn.cursor()
    row = cursor.execute("SELECT is_authorized FROM users WHERE username = ?", (username,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found.")
        
    new_status = False if row["is_authorized"] else True
    cursor.execute("UPDATE users SET is_authorized = ? WHERE username = ?", (new_status, username))
    conn.commit()
    conn.close()
    return {"success": True, "username": username, "is_authorized": bool(new_status)}

# Attachments Management API
ATTACHMENTS_DIR = os.path.join("uploads", "attachments")
os.makedirs(ATTACHMENTS_DIR, exist_ok=True)

@app.get("/api/attachments")
def list_attachments():
    """List all custom files uploaded to the attachments folder."""
    files = []
    if os.path.exists(ATTACHMENTS_DIR):
        for f in os.listdir(ATTACHMENTS_DIR):
            file_path = os.path.join(ATTACHMENTS_DIR, f)
            if os.path.isfile(file_path) and f != ".gitkeep":
                files.append({
                    "filename": f,
                    "size": os.path.getsize(file_path)
                })
    return files

@app.post("/api/attachments/upload")
def upload_attachment(file: UploadFile = File(...)):
    """Upload a custom attachment file to be sent in the student emails."""
    try:
        # Secure the filename
        safe_filename = file.filename.replace(" ", "_")
        dest_path = os.path.join(ATTACHMENTS_DIR, safe_filename)
        with open(dest_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"success": True, "filename": safe_filename, "message": "Attachment uploaded successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload attachment: {e}")

@app.delete("/api/attachments/{filename}")
def delete_attachment(filename: str):
    """Delete a custom attachment file."""
    safe_filename = os.path.basename(filename)
    file_path = os.path.join(ATTACHMENTS_DIR, safe_filename)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            return {"success": True, "message": f"Attachment {safe_filename} deleted."}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to delete attachment: {e}")
    raise HTTPException(status_code=404, detail="Attachment file not found.")

@app.get("/api/system/health")
def get_system_health():
    """Check backend database connection and status."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        conn.close()
        return {
            "status": "online",
            "database": "postgresql" if config.IS_POSTGRES else "sqlite",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "status": "offline",
            "error": str(e),
            "database": "postgresql" if config.IS_POSTGRES else "sqlite",
            "timestamp": datetime.now().isoformat()
        }

@app.get("/api/system/integrity")
def get_system_integrity():
    """Run full diagnostic checks on the database records and file mappings."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        students = cursor.execute("SELECT student_id, name, email, department FROM students").fetchall()
        
        total_students = len(students)
        qr_exists_count = 0
        missing_qrcodes = []
        missing_emails = []
        missing_departments = []
        
        # Get stats counts
        emails_sent = cursor.execute("SELECT COUNT(*) FROM students WHERE email_sent = TRUE").fetchone()[0]
        checked_in = cursor.execute("SELECT COUNT(*) FROM students WHERE checked_in = TRUE").fetchone()[0]
        
        # Check duplicate student IDs
        duplicates_query = cursor.execute("""
            SELECT student_id, COUNT(*) as cnt 
            FROM students 
            GROUP BY student_id 
            HAVING COUNT(*) > 1
        """).fetchall()
        duplicate_student_ids = [{"student_id": d["student_id"], "count": d["cnt"]} for d in duplicates_query]
        
        # Check files and empty fields
        for s in students:
            s_id = s["student_id"]
            qr_path = os.path.join(config.QRCODES_DIR, f"{s_id}.png")
            if os.path.exists(qr_path):
                qr_exists_count += 1
            else:
                missing_qrcodes.append({
                    "student_id": s_id,
                    "name": s["name"],
                    "issue": "Missing QR Code file"
                })
                
            email = s["email"]
            if not email or email.strip() == "" or "@" not in email:
                missing_emails.append({
                    "student_id": s_id,
                    "name": s["name"],
                    "email": email or "",
                    "issue": "Missing or invalid email"
                })
                
            dept = s["department"]
            if not dept or dept.strip() == "" or dept.lower() == "nan":
                missing_departments.append({
                    "student_id": s_id,
                    "name": s["name"],
                    "issue": "Missing department"
                })
                
        conn.close()
        
        return {
            "db_status": "online",
            "db_type": "postgresql" if config.IS_POSTGRES else "sqlite",
            "total_students": total_students,
            "total_qr_generated": qr_exists_count,
            "total_emails_sent": emails_sent,
            "total_attendance": checked_in,
            "duplicates": duplicate_student_ids,
            "missing_qrcodes": missing_qrcodes,
            "missing_emails": missing_emails,
            "missing_departments": missing_departments,
            "last_sync_time": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to run integrity check: {str(e)}")

@app.post("/api/system/integrity/check")
def run_integrity_check_trigger():
    """Triggers and returns validation status for the event readiness."""
    res = get_system_integrity()
    
    total_issues = (
        len(res["duplicates"]) + 
        len(res["missing_qrcodes"]) + 
        len(res["missing_emails"]) + 
        len(res["missing_departments"])
    )
    
    if total_issues == 0 and res["total_students"] > 0:
        return {
            "ready": True,
            "message": "System Ready for Event",
            "details": res
        }
    else:
        issues_summary = []
        if res["total_students"] == 0:
            issues_summary.append("Database is empty. Please upload student roster.")
        if len(res["duplicates"]) > 0:
            issues_summary.append(f"Found {len(res['duplicates'])} duplicate student IDs.")
        if len(res["missing_qrcodes"]) > 0:
            issues_summary.append(f"Found {len(res['missing_qrcodes'])} missing QR code pass files.")
        if len(res["missing_emails"]) > 0:
            issues_summary.append(f"Found {len(res['missing_emails'])} students with missing/invalid emails.")
        if len(res["missing_departments"]) > 0:
            issues_summary.append(f"Found {len(res['missing_departments'])} students with missing departments.")
            
        return {
            "ready": False,
            "message": "Integrity check failed. Please resolve the listed issues before the event.",
            "issues_summary": issues_summary,
            "details": res
        }
